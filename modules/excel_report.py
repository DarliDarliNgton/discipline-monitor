"""
Генератор Excel-отчёта о нарушениях.
"""

from __future__ import annotations

import io
import time
from typing import List

from modules.logger import get_logger
from modules.violation_tracker import Violation

log = get_logger(__name__)

_RU = {
    "sleeping":    "Сон на занятии",
    "phone_usage": "Использование телефона",
    "bottle":      "Запрещённый предмет (бутылка)",
    "food":        "Запрещённый предмет (еда)",
}


def generate_excel(violations: List[Violation], session_start: float) -> bytes:
    """Возвращает .xlsx файл как bytes для скачивания через Streamlit."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Журнал нарушений"

        # Цвета
        HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
        EVEN_FILL   = PatternFill("solid", fgColor="F0F0FF")
        RED_FILL    = PatternFill("solid", fgColor="FFDDDD")
        ORANGE_FILL = PatternFill("solid", fgColor="FFE8CC")

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Заголовок
        ws.merge_cells("A1:G1")
        ws["A1"] = "ОТЧЁТ О НАРУШЕНИЯХ ДИСЦИПЛИНЫ"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = HEADER_FILL
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Мета-информация
        date_str = datetime.datetime.fromtimestamp(session_start).strftime("%Y-%m-%d")
        start_str = datetime.datetime.fromtimestamp(session_start).strftime("%H:%M:%S")
        end_str = datetime.datetime.now().strftime("%H:%M:%S")
        ws["A2"] = f"Дата: {date_str}  |  Время мониторинга: {start_str} – {end_str}  |  Всего нарушений: {len(violations)}"
        ws["A2"].font = Font(italic=True, color="666666")
        ws.merge_cells("A2:G2")

        # Шапка таблицы
        headers = ["#", "Тип нарушения", "Время начала", "Длительность (с)",
                   "Нарушитель", "Уверенность", "Видеозапись"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Данные
        color_map = {
            "sleeping":    RED_FILL,
            "phone_usage": ORANGE_FILL,
        }
        for i, v in enumerate(violations, 1):
            row = 4 + i
            fill = color_map.get(v.cls_name, EVEN_FILL) if i % 2 == 0 else PatternFill()
            start_fmt = datetime.datetime.fromtimestamp(v.start_time).strftime("%H:%M:%S")
            values = [
                i,
                _RU.get(v.cls_name, v.cls_name),
                start_fmt,
                int(v.duration),
                v.person_name,
                f"{v.confidence:.0%}",
                v.video_segment_path or "—",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
                if fill.fill_type:
                    cell.fill = fill

        # Статистика
        stat_row = 4 + len(violations) + 2
        ws.cell(row=stat_row, column=1, value="Итого по типам:").font = Font(bold=True)
        by_type: dict[str, int] = {}
        for v in violations:
            by_type[v.cls_name] = by_type.get(v.cls_name, 0) + 1
        for j, (cls, cnt) in enumerate(by_type.items(), 1):
            ws.cell(row=stat_row + j, column=1, value=_RU.get(cls, cls))
            ws.cell(row=stat_row + j, column=2, value=cnt)

        # Ширина столбцов
        widths = [5, 30, 14, 18, 20, 12, 40]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        log.info("Excel отчёт сгенерирован: %d нарушений", len(violations))
        return buf.getvalue()
    except ImportError:
        log.error("openpyxl не установлен")
        raise RuntimeError("Установите openpyxl: pip install openpyxl")
    except Exception as e:
        log.exception("Ошибка генерации Excel: %s", e)
        raise
